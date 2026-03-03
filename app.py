"""
=============================================================
  AUDITORIA DE REFEIÇÕES - SISTERMI  v1.2
  Execute: pip install -r requirements.txt
           streamlit run app.py
=============================================================
"""
import io, re
from datetime import date, datetime
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Auditoria de Refeições", page_icon="🍽️",
                   layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.stApp{background:#0f1117;color:#e8eaf0;}
.audit-header{background:linear-gradient(135deg,#1a1f2e,#0d1117);border:1px solid #2a3040;
  border-radius:16px;padding:36px 40px;margin-bottom:28px;}
.audit-header h1{font-size:2rem;font-weight:700;color:#f0f4ff;margin:0 0 6px 0;}
.audit-header p{color:#6b7280;font-size:.95rem;margin:0;}
.audit-badge{display:inline-block;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.3);
  color:#60a5fa;font-size:.75rem;font-weight:600;padding:3px 10px;border-radius:20px;
  margin-bottom:12px;letter-spacing:1px;text-transform:uppercase;}
.metric-card{background:#13161f;border:1px solid #2a3040;border-radius:12px;padding:20px 24px;text-align:center;}
.metric-card .value{font-family:'DM Mono',monospace;font-size:2.4rem;font-weight:600;line-height:1;margin-bottom:6px;}
.metric-card .label{font-size:.8rem;color:#6b7280;text-transform:uppercase;letter-spacing:.8px;}
.color-green{color:#34d399;}.color-yellow{color:#fbbf24;}.color-red{color:#f87171;}.color-blue{color:#60a5fa;}
.alert-inc{background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.25);
  border-radius:10px;padding:16px 20px;margin-bottom:10px;}
.alert-inc .mat{font-family:'DM Mono',monospace;color:#9ca3af;font-size:.8rem;}
.alert-inc .nome{font-weight:600;color:#f0f4ff;}
.alert-inc .motivo{color:#f87171;font-size:.85rem;margin-top:4px;}
.section-title{font-size:1rem;font-weight:600;color:#d1d5db;margin:28px 0 14px 0;
  padding-bottom:8px;border-bottom:1px solid #2a3040;}
.stDownloadButton>button{background:linear-gradient(135deg,#3b82f6,#2563eb)!important;
  color:white!important;border:none!important;border-radius:10px!important;
  padding:12px 28px!important;font-weight:600!important;width:100%!important;}
footer{visibility:hidden;}#MainMenu{visibility:hidden;}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ────────────────────────────────────────────────────────────────
STATUS_KW = ["Compensado","DSR","Férias","Sem Marcação","Lic.Medica",
             "Auxílio","Folga","Afastamento","Redução","Aviso"]
DAY_RE = re.compile(r"^(\d{1,2})\s+\w{3}\s+(\d{4})\s+(.+)$")
TIME_RE = re.compile(r"\b(\d{2}:\d{2})\b")
EMP_RE  = re.compile(r"Empregado:\s*(\d+)\s+(.+?)\s+Escala")
DIAS_SEM = {0:"SEG",1:"TER",2:"QUA",3:"QUI",4:"SEX",5:"SAB",6:"DOM"}
ALMOCO_WIN = (11*60, 17*60)
JANTA_WIN  = (19*60, 23*60+59)

# ─── NORMALIZACAO ──────────────────────────────────────────────────────────────
# Usa ALMOCO (sem cedilha) como padrão interno para evitar bugs de encoding Windows
def norm_tipo(t: str) -> str:
    return str(t).strip().upper().replace("ALMOÇO","ALMOCO")

def get_col(df, *kws):
    """Detecta coluna por keyword - robusto a variações de nome"""
    for kw in kws:
        for c in df.columns:
            if kw.upper() in str(c).upper():
                return c
    return None

# ─── PARSER PDF ────────────────────────────────────────────────────────────────
def _get_entrada_saida(resto):
    t = TIME_RE.findall(resto)
    if not t: return None, None
    if len(t) == 1: return t[0], None
    h, _ = map(int, t[-1].split(":"))
    return (t[0], t[-2]) if h < 13 else (t[0], t[-1])

def parse_pdf(pdf_bytes: bytes) -> pd.DataFrame:
    records = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            lines = text.split("\n")
            mat = nome = p_ini = None
            for line in lines:
                m = EMP_RE.search(line)
                if m: mat = m.group(1).strip(); nome = m.group(2).strip()
                p = re.search(r"Período:\s*(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})", line)
                if p: p_ini = datetime.strptime(p.group(1), "%d/%m/%Y").date()
            if not mat: continue
            in_d = False
            for line in lines:
                if "DT Sem Hor" in line: in_d = True; continue
                if any(x in line for x in ["TOTAIS","Saldo de Banco","Estou de pleno"]):
                    in_d = False; continue
                if not in_d: continue
                m = DAY_RE.match(line.strip())
                if not m: continue
                d_num = int(m.group(1)); resto = m.group(3).strip()
                if not p_ini: continue
                try:
                    if d_num >= p_ini.day: dr = date(p_ini.year, p_ini.month, d_num)
                    else:
                        mn = p_ini.month % 12 + 1
                        an = p_ini.year + (1 if p_ini.month == 12 else 0)
                        dr = date(an, mn, d_num)
                except ValueError: continue
                if any(k in resto for k in STATUS_KW): st = "NAO_TRABALHADO"; e = s = None
                elif TIME_RE.findall(resto): st = "TRABALHADO"; e, s = _get_entrada_saida(resto)
                else: st = "NAO_TRABALHADO"; e = s = None
                records.append({"matricula":mat,"nome_pdf":nome,"data":dr,"entrada":e,"saida":s,"status":st})
    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["matricula","nome_pdf","data","entrada","saida","status"])

# ─── PARSER EXCEL REFEIÇÕES ────────────────────────────────────────────────────
def parse_excel_refeicoes(xlsx_bytes: bytes):
    # Lê pela POSIÇÃO da aba (índice 0), não pelo nome — evita problema de encoding no Windows
    df_raw = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=0, header=None)

    try: restaurante = str(df_raw.iloc[1][2])
    except: restaurante = "N/D"

    linha0 = df_raw.iloc[0].tolist()
    def to_date(v, ano_ref=None):
        if isinstance(v, datetime):
            d = v.date()
            # Corrige ano errado (ex: planilha com 2025 onde deveria ser 2026)
            if ano_ref and d.year != ano_ref:
                d = date(ano_ref, d.month, d.day)
            return d
        try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        except: return None

    periodo_ini = to_date(linha0[8])
    periodo_fim = to_date(linha0[12], ano_ref=periodo_ini.year if periodo_ini else None)

    df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=0, header=2)
    cols = ["_","MATRICULA","NOME","TIPO"] + [f"DIA_{i}" for i in range(1,32)] + ["TOTAL"]
    df.columns = cols[:len(df.columns)]
    df = df[df["MATRICULA"].notna()].copy()
    df["MATRICULA"] = df["MATRICULA"].astype(float).astype(int).astype(str)
    df["TIPO"] = df["TIPO"].apply(norm_tipo)  # normaliza sem cedilha

    meals = []
    for _, row in df.iterrows():
        for i in range(1, 32):
            if pd.notna(row.get(f"DIA_{i}")) and row[f"DIA_{i}"] == 1:
                ano = periodo_ini.year if periodo_ini else 2026
                mes = periodo_ini.month if periodo_ini else 2
                try:
                    meals.append({"matricula": str(row["MATRICULA"]), "nome": row["NOME"],
                                   "tipo": row["TIPO"], "data": date(ano, mes, i)})
                except ValueError: pass

    return pd.DataFrame(meals), periodo_ini, periodo_fim, restaurante

# ─── PARSER EXCEL EXCEÇÕES ────────────────────────────────────────────────────
def parse_excel_excecoes(xlsx_bytes: bytes) -> pd.DataFrame:
    COLS = ["matricula","data_inicio","data_fim","tipos","autorizado","motivo"]

    # Lê pela POSIÇÃO (índice 1 = segunda aba), nunca pelo nome da aba
    try:
        df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=1)
    except Exception:
        return pd.DataFrame(columns=COLS)

    if df.empty:
        return pd.DataFrame(columns=COLS)

    mat_col     = get_col(df, "MATRI")
    col_datafim = get_col(df, "DATA_FIM", "DATAFIM", "FIM")
    col_tipo    = get_col(df, "TIPO")
    col_motivo  = get_col(df, "MOTIVO")
    col_aut     = get_col(df, "AUTORIZADO", "S/N")

    if not mat_col:
        return pd.DataFrame(columns=COLS)

    # Converte matrícula de forma segura (int64, float, string — tudo funciona)
    df = df[df[mat_col].notna()].copy()
    df["_mat"] = df[mat_col].apply(lambda x: str(int(float(x))))

    records = []
    for _, row in df.iterrows():
        # Período da exceção — vem como "01/02/2026 ate 13/02/2026"
        datafim_raw = str(row[col_datafim]) if col_datafim else ""
        ms = re.findall(r"(\d{2}/\d{2}/\d{4})", datafim_raw)
        ini = datetime.strptime(ms[0], "%d/%m/%Y").date() if len(ms) >= 1 else date(2000,1,1)
        fim = datetime.strptime(ms[1], "%d/%m/%Y").date() if len(ms) >= 2 else date(2099,12,31)

        # Tipo — normalizado igual às refeições (sem cedilha)
        tp_raw = norm_tipo(str(row[col_tipo]) if col_tipo else "")
        tipos: set = set()
        if "ALMOCO" in tp_raw: tipos.add("ALMOCO")
        if "JANTA"  in tp_raw: tipos.add("JANTA")
        if not tipos: tipos = {"ALMOCO", "JANTA"}  # sem tipo = vale para ambos

        # Autorizado S/N
        aut = str(row[col_aut]).strip().upper() if col_aut else "S"
        if aut not in ("S","N"): aut = "S"

        motivo = str(row[col_motivo]).strip() if col_motivo else ""

        records.append({"matricula": row["_mat"], "data_inicio": ini, "data_fim": fim,
                         "tipos": tipos, "autorizado": aut, "motivo": motivo})

    return pd.DataFrame(records, columns=COLS) if records else pd.DataFrame(columns=COLS)

# ─── MOTOR DE CONFORMIDADE ─────────────────────────────────────────────────────
def _to_min(t): h, m = map(int, t.split(":")); return h*60+m

def classify(tipo_norm, entrada, saida, status):
    if status != "TRABALHADO":
        return "INCONFORME", "DIA NÃO TRABALHADO (DSR/FOLGA/FÉRIAS/LIC)"
    if not entrada or not saida:
        return "INCONFORME", "SEM JORNADA NO CARTÃO"
    e, s = _to_min(entrada), _to_min(saida)
    win = ALMOCO_WIN if "ALMOCO" in tipo_norm else JANTA_WIN
    if max(0, min(s, win[1]) - max(e, win[0])) >= 1:
        return "CONFORME", "OK"
    return "INCONFORME", f"FORA DA JANELA DO {tipo_norm}"

def run_audit(df_meals, df_shifts, df_exc):
    exc_ok = not df_exc.empty and "matricula" in df_exc.columns
    results = []

    for _, meal in df_meals.iterrows():
        mat  = meal["matricula"]
        data = meal["data"]
        tipo = meal["tipo"]   # já normalizado: ALMOCO ou JANTA

        # Cartão de ponto
        sh = df_shifts[(df_shifts["matricula"] == mat) & (df_shifts["data"] == data)]
        if sh.empty:
            rb, mb = "INCONFORME", "SEM CARTÃO DE PONTO NO PDF"
            e = s = None; sp = "SEM_REGISTRO"
        else:
            r = sh.iloc[0]
            e, s, sp = r["entrada"], r["saida"], r["status"]
            rb, mb = classify(tipo, e, s, sp)

        # Exceção — tipos já normalizados dos dois lados (ALMOCO/JANTA)
        res = rb; mot_exc = ""; flag_exc = False
        if exc_ok:
            cands = df_exc[
                (df_exc["matricula"]   == mat) &
                (df_exc["data_inicio"] <= data) &
                (df_exc["data_fim"]    >= data)
            ]
            cands = cands[cands["tipos"].apply(lambda x: tipo in x)]

            if not cands.empty:
                # N tem prioridade; entre múltiplos S, pega o primeiro
                cn = cands[cands["autorizado"] == "N"]
                ch = cn.iloc[0] if not cn.empty else cands[cands["autorizado"] == "S"].iloc[0]
                res      = "CONFORME (EXCEÇÃO)" if ch["autorizado"] == "S" else "INCONFORME (EXCEÇÃO)"
                mot_exc  = ch["motivo"]
                flag_exc = True

        # Converte tipo de volta para exibição com acento
        tipo_display = "ALMOÇO" if "ALMOCO" in tipo else tipo

        results.append({
            "matricula":      mat,
            "nome":           meal["nome"],
            "data":           data,
            "dia_sem":        DIAS_SEM.get(data.weekday(), ""),
            "tipo":           tipo_display,
            "entrada":        e,
            "saida":          s,
            "status_ponto":   sp,
            "resultado_base": rb,
            "motivo_base":    mb,
            "resultado":      res,
            "motivo_excecao": mot_exc,
            "flag_excecao":   flag_exc,
        })
    return pd.DataFrame(results)

# ─── GERADOR EXCEL RELATÓRIO ───────────────────────────────────────────────────
COR_HDR="1A2744"; COR_CONF="C6EFCE"; COR_EXC="FFEB9C"
COR_INC="FFC7CE"; COR_TOT="D9E1F2"; BRANCO="FFFFFF"

def _tb():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg=COR_HDR):
    cell.font      = Font(bold=True, color=BRANCO, name="Calibri", size=9)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _tb()

def _c(cell, v, cor=None, bold=False, fmt=None, align="center"):
    cell.value     = v
    cell.font      = Font(name="Calibri", size=9, bold=bold)
    cell.alignment = Alignment(vertical="center", horizontal=align, wrap_text=True)
    cell.border    = _tb()
    if cor:  cell.fill          = PatternFill("solid", start_color=cor)
    if fmt:  cell.number_format = fmt

def gerar_excel(df_res, restaurante, periodo_ini, periodo_fim):
    wb = Workbook()
    ps = (f"{periodo_ini.strftime('%d/%m/%Y')} a {periodo_fim.strftime('%d/%m/%Y')}"
          if periodo_ini and periodo_fim else "")

    # ── ABA 1: RESUMO ─────────────────────────────────────────
    ws = wb.active; ws.title = "RESUMO"; ws.sheet_view.showGridLines = False

    def titulo(ws, txt, span, bg, size=12):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value = txt
        c.font  = Font(bold=True, color=BRANCO, name="Calibri", size=size)
        c.fill  = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    titulo(ws, "RELATÓRIO DE AUDITORIA DE REFEIÇÕES", "A1:I1", COR_HDR, 14)
    ws.row_dimensions[1].height = 36
    titulo(ws, f"Restaurante: {restaurante}   |   Período: {ps}", "A2:I2", "2E4B87", 10)
    ws.row_dimensions[2].height = 20

    total = len(df_res)
    nc  = int((df_res["resultado"] == "CONFORME").sum())
    nce = int((df_res["resultado"] == "CONFORME (EXCEÇÃO)").sum())
    ni  = int((df_res["resultado"] == "INCONFORME").sum())
    nie = int((df_res["resultado"] == "INCONFORME (EXCEÇÃO)").sum())
    nok = nc + nce
    pct = nok / total if total else 0

    kpi = [("TOTAL",total,None,"D9E1F2"),("CONFORMES",nc,None,"C6EFCE"),
           ("CONF.(EXCEÇÃO)",nce,None,"FFEB9C"),("INCONFORMES",ni,None,"FFC7CE"),
           ("INC.(EXCEÇÃO)",nie,None,"FFCCCC"),("TOTAL CONF.",nok,None,"BDD7EE"),
           ("CONFORMIDADE",pct,"0.0%","BDD7EE")]
    ws.row_dimensions[4].height = 34
    ws.row_dimensions[5].height = 28
    for ci,(lbl,val,fmt,cor) in enumerate(kpi, 1):
        cl = ws.cell(row=4, column=ci, value=lbl)
        cl.font=Font(bold=True,name="Calibri",size=8); cl.fill=PatternFill("solid",start_color=cor)
        cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cl.border=_tb()
        cv = ws.cell(row=5, column=ci, value=val)
        cv.font=Font(bold=True,name="Calibri",size=14); cv.fill=PatternFill("solid",start_color=cor)
        cv.alignment=Alignment(horizontal="center",vertical="center"); cv.border=_tb()
        if fmt: cv.number_format = fmt

    ws.row_dimensions[7].height = 24
    for ci,h in enumerate(["MATRÍCULA","NOME","TOTAL","CONFORMES","CONF.(EXC.)",
                            "INCONFORMES","INC.(EXC.)","TOTAL CONF.","% CONF."], 1):
        _hdr(ws.cell(row=7, column=ci)); ws.cell(row=7, column=ci).value = h

    row = 8
    for mat in df_res["matricula"].unique():
        sub = df_res[df_res["matricula"] == mat]
        nome = sub.iloc[0]["nome"]; t_ = len(sub)
        cf = int((sub["resultado"]=="CONFORME").sum())
        ce = int((sub["resultado"]=="CONFORME (EXCEÇÃO)").sum())
        ic = int((sub["resultado"]=="INCONFORME").sum())
        ie = int((sub["resultado"]=="INCONFORME (EXCEÇÃO)").sum())
        tok = cf + ce; pct_ = tok / t_ if t_ else 0
        cor_r = "FFF2CC" if ic + ie > 0 else None
        for ci,v in enumerate([mat,nome,t_,cf,ce,ic,ie,tok,pct_], 1):
            _c(ws.cell(row=row, column=ci), v, cor=cor_r,
               align="left" if ci==2 else "center",
               fmt="0.0%" if ci==9 else None)
        row += 1

    # linha de totais
    for ci in range(1, 10):
        _c(ws.cell(row=row, column=ci), "", cor=COR_TOT, bold=True)
    _c(ws.cell(row=row,column=1), "TOTAL GERAL", cor=COR_TOT, bold=True)
    for ci,col in zip(range(3,9), ["C","D","E","F","G","H"]):
        c = ws.cell(row=row, column=ci)
        c.value = f"=SUM({col}8:{col}{row-1})"
        c.font=Font(bold=True,name="Calibri",size=9); c.fill=PatternFill("solid",start_color=COR_TOT)
        c.alignment=Alignment(horizontal="center"); c.border=_tb()
    c9 = ws.cell(row=row, column=9)
    c9.value=f"=H{row}/C{row}"; c9.number_format="0.0%"
    c9.font=Font(bold=True,name="Calibri",size=9); c9.fill=PatternFill("solid",start_color=COR_TOT)
    c9.alignment=Alignment(horizontal="center"); c9.border=_tb()

    for ci,w in zip(range(1,10), [12,38,8,11,13,12,11,12,10]):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── ABA 2: DETALHE ─────────────────────────────────────────
    ws2 = wb.create_sheet("DETALHE"); ws2.sheet_view.showGridLines = False
    titulo(ws2, "DETALHE COMPLETO – TODAS AS REFEIÇÕES", "A1:K1", COR_HDR, 11)
    ws2.row_dimensions[1].height = 24
    for ci,h in enumerate(["MATRÍCULA","NOME","DATA","DIA","TIPO","ENTRADA","SAÍDA",
                            "STATUS PONTO","RESULTADO BASE","RESULTADO FINAL","MOTIVO / EXCEÇÃO"], 1):
        _hdr(ws2.cell(row=2, column=ci)); ws2.cell(row=2, column=ci).value = h
    ws2.row_dimensions[2].height = 26

    for ri, (_, r) in enumerate(df_res.iterrows(), 3):
        res = r["resultado"]
        cor = COR_CONF if res=="CONFORME" else (COR_EXC if "EXCEÇÃO" in res else COR_INC)
        # Motivo: se tem exceção mostra o motivo da exceção, senão mostra o motivo base
        motivo_col = r["motivo_excecao"] if r["flag_excecao"] else r["motivo_base"]
        vals = [r["matricula"], r["nome"], r["data"].strftime("%d/%m/%Y"), r["dia_sem"],
                r["tipo"], r["entrada"] or "—", r["saida"] or "—",
                r["status_ponto"], r["resultado_base"], r["resultado"], motivo_col]
        for ci,v in enumerate(vals, 1):
            _c(ws2.cell(row=ri, column=ci), v, cor=cor,
               align="left" if ci in [2,11] else "center")

    for ci,w in zip(range(1,12), [12,36,12,6,8,8,8,18,28,22,42]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── ABA 3: INCONFORMES ─────────────────────────────────────
    ws3 = wb.create_sheet("INCONFORMES"); ws3.sheet_view.showGridLines = False
    titulo(ws3, "⚠  INCONFORMES — SEM JUSTIFICATIVA", "A1:H1", "C00000", 11)
    ws3.row_dimensions[1].height = 24
    for ci,h in enumerate(["MATRÍCULA","NOME","DATA","DIA","TIPO","ENTRADA","SAÍDA","MOTIVO"], 1):
        _hdr(ws3.cell(row=2, column=ci), bg="C00000")
        ws3.cell(row=2, column=ci).value = h

    inc_df = df_res[df_res["resultado"] == "INCONFORME"]
    if inc_df.empty:
        ws3.merge_cells("A3:H3")
        ws3["A3"] = "✓  Nenhum inconforme puro neste período."
        ws3["A3"].font = Font(name="Calibri", size=10, color="375623")
        ws3["A3"].fill = PatternFill("solid", start_color=COR_CONF)
        ws3["A3"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        for ri,(_, r) in enumerate(inc_df.iterrows(), 3):
            for ci,v in enumerate([r["matricula"], r["nome"], r["data"].strftime("%d/%m/%Y"),
                                    r["dia_sem"], r["tipo"], r["entrada"] or "—",
                                    r["saida"] or "—", r["motivo_base"]], 1):
                _c(ws3.cell(row=ri, column=ci), v, cor=COR_INC,
                   align="left" if ci in [2,8] else "center")

    for ci,w in zip(range(1,9), [12,36,12,6,8,8,8,40]):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── ABA 4: EXCEÇÕES APLICADAS ──────────────────────────────
    ws4 = wb.create_sheet("EXCEÇÕES APLICADAS"); ws4.sheet_view.showGridLines = False
    titulo(ws4, "EXCEÇÕES AUTORIZADAS / APLICADAS", "A1:I1", "7B5B00", 11)
    ws4.row_dimensions[1].height = 24
    for ci,h in enumerate(["MATRÍCULA","NOME","DATA","DIA","TIPO","SEM PONTO?",
                            "RESULTADO S/ EXC.","RESULTADO FINAL","MOTIVO DA EXCEÇÃO"], 1):
        _hdr(ws4.cell(row=2, column=ci), bg="7B5B00")
        ws4.cell(row=2, column=ci).value = h

    exc_df = df_res[df_res["flag_excecao"] == True]
    if exc_df.empty:
        ws4.merge_cells("A3:I3")
        ws4["A3"] = "Nenhuma exceção aplicada neste período."
        ws4["A3"].alignment = Alignment(horizontal="center")
    else:
        for ri,(_, r) in enumerate(exc_df.iterrows(), 3):
            sem_ponto = "⚠ SIM" if r["status_ponto"] == "SEM_REGISTRO" else "NÃO"
            for ci,v in enumerate([r["matricula"], r["nome"], r["data"].strftime("%d/%m/%Y"),
                                    r["dia_sem"], r["tipo"], sem_ponto,
                                    r["resultado_base"], r["resultado"], r["motivo_excecao"]], 1):
                _c(ws4.cell(row=ri, column=ci), v, cor=COR_EXC,
                   align="left" if ci in [2,9] else "center")

    for ci,w in zip(range(1,10), [12,36,12,6,8,10,28,22,42]):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ─── UI PRINCIPAL ──────────────────────────────────────────────────────────────
def main():
    st.markdown("""
    <div class="audit-header">
        <div class="audit-badge">SISTERMI</div>
        <h1>🍽️ Auditoria de Refeições</h1>
        <p>Upload do Excel de refeições + PDF de ponto → relatório completo de conformidade em segundos.</p>
    </div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### 📊 Planilha de Refeições (.xlsx)")
        st.caption("Abas: REFEIÇÕES · EXCEÇÕES")
        file_xlsx = st.file_uploader("xlsx", type=["xlsx"], key="xu", label_visibility="collapsed")
        if file_xlsx: st.success(f"✓ {file_xlsx.name}  ({file_xlsx.size//1024} KB)")
    with col2:
        st.markdown("##### 📋 Espelho de Ponto (.pdf)")
        st.caption("PDF com cartões de ponto dos colaboradores")
        file_pdf = st.file_uploader("pdf", type=["pdf"], key="pu", label_visibility="collapsed")
        if file_pdf: st.success(f"✓ {file_pdf.name}  ({file_pdf.size//1024} KB)")

    st.markdown("<br>", unsafe_allow_html=True)
    btn_col, _ = st.columns([1, 3])
    with btn_col:
        run_btn = st.button("▶  Gerar Auditoria", type="primary",
                            disabled=not(file_xlsx and file_pdf), use_container_width=True)

    if not (file_xlsx and file_pdf):
        st.info("⬆  Faça upload dos dois arquivos para habilitar a auditoria.")
        return
    if not run_btn and "df_res" not in st.session_state:
        return

    if run_btn:
        with st.spinner("⚙  Processando arquivos..."):
            xlsx_bytes = file_xlsx.read()
            pdf_bytes  = file_pdf.read()
            df_meals, periodo_ini, periodo_fim, restaurante = parse_excel_refeicoes(xlsx_bytes)
            df_shifts = parse_pdf(pdf_bytes)
            df_exc    = parse_excel_excecoes(xlsx_bytes)
            df_res    = run_audit(df_meals, df_shifts, df_exc)
            st.session_state.update({"df_res": df_res, "restaurante": restaurante,
                                      "periodo_ini": periodo_ini, "periodo_fim": periodo_fim})

    df_res      = st.session_state.get("df_res")
    restaurante = st.session_state.get("restaurante", "")
    periodo_ini = st.session_state.get("periodo_ini")
    periodo_fim = st.session_state.get("periodo_fim")

    if df_res is None or df_res.empty:
        st.error("Nenhum dado processado. Verifique os arquivos."); return

    # ── KPIs
    total = len(df_res)
    nc    = int((df_res["resultado"] == "CONFORME").sum())
    nce   = int((df_res["resultado"] == "CONFORME (EXCEÇÃO)").sum())
    ni    = int((df_res["resultado"] == "INCONFORME").sum())
    nok   = nc + nce
    pct   = f"{nok/total*100:.1f}%" if total else "0%"

    st.markdown(f"""
    <div class="section-title">📈 Resumo &nbsp;
        <span style="color:#6b7280;font-size:.85rem;font-weight:400">
        {restaurante} &nbsp;|&nbsp;
        {periodo_ini.strftime('%d/%m/%Y') if periodo_ini else '—'} a
        {periodo_fim.strftime('%d/%m/%Y') if periodo_fim else '—'}
        </span>
    </div>""", unsafe_allow_html=True)

    for col, val, lbl, cls in zip(st.columns(5),
        [total, nc, nce, ni, pct],
        ["TOTAL REFEIÇÕES","CONFORMES","CONF. (EXCEÇÃO)","INCONFORMES","CONFORMIDADE"],
        ["color-blue","color-green","color-yellow","color-red","color-blue"]):
        with col:
            st.markdown(f"""<div class="metric-card">
                <div class="value {cls}">{val}</div>
                <div class="label">{lbl}</div></div>""", unsafe_allow_html=True)

    # ── INCONFORMES (sem exceção)
    inc_df = df_res[df_res["resultado"] == "INCONFORME"]
    if not inc_df.empty:
        st.markdown(f"""
        <div class="section-title">⚠️ Inconformes sem Justificativa &nbsp;
            <span style="background:rgba(248,113,113,.15);color:#f87171;font-size:.75rem;
                  padding:2px 8px;border-radius:12px;font-weight:600">
            {len(inc_df)} ocorrências</span></div>""", unsafe_allow_html=True)
        for _, r in inc_df.iterrows():
            st.markdown(f"""<div class="alert-inc">
                <span class="mat">{r['matricula']}</span>
                <span class="nome"> — {r['nome']}</span>
                <div class="motivo">
                📅 {r['data'].strftime('%d/%m/%Y')} ({r['dia_sem']}) &nbsp;|&nbsp;
                🍽 {r['tipo']} &nbsp;|&nbsp;
                ⏰ Entrada: {r['entrada'] or '—'} / Saída: {r['saida'] or '—'} &nbsp;|&nbsp;
                ❌ {r['motivo_base']}</div></div>""", unsafe_allow_html=True)
    else:
        st.success("✅ Nenhum inconforme sem justificativa neste período.")

    # ── EXCEÇÕES APLICADAS (pessoas sem cartão ou fora da janela mas com exceção)
    exc_app = df_res[df_res["flag_excecao"] == True]
    if not exc_app.empty:
        st.markdown(f"""
        <div class="section-title">📋 Exceções Aplicadas &nbsp;
            <span style="background:rgba(251,191,36,.15);color:#fbbf24;font-size:.75rem;
                  padding:2px 8px;border-radius:12px;font-weight:600">
            {len(exc_app)} refeições</span></div>""", unsafe_allow_html=True)
        df_show = exc_app[["matricula","nome","data","tipo","status_ponto","resultado_base","resultado","motivo_excecao"]].copy()
        df_show["data"] = df_show["data"].apply(lambda d: d.strftime("%d/%m/%Y"))
        df_show["Sem Ponto?"] = df_show["status_ponto"].apply(lambda x: "⚠ SIM" if x=="SEM_REGISTRO" else "NÃO")
        df_show = df_show.rename(columns={
            "matricula":"Matríc.","nome":"Nome","data":"Data","tipo":"Tipo",
            "resultado_base":"Resultado S/ Exc.","resultado":"Resultado Final","motivo_excecao":"Motivo Exceção"
        })[["Matríc.","Nome","Data","Tipo","Sem Ponto?","Resultado S/ Exc.","Resultado Final","Motivo Exceção"]]
        st.dataframe(df_show, use_container_width=True, hide_index=True)

    # ── CONSOLIDADO
    st.markdown('<div class="section-title">📊 Consolidado por Colaborador</div>', unsafe_allow_html=True)
    summary = []
    for mat in df_res["matricula"].unique():
        sub = df_res[df_res["matricula"] == mat]; nome = sub.iloc[0]["nome"]; t_ = len(sub)
        cf  = int((sub["resultado"]=="CONFORME").sum())
        ce  = int((sub["resultado"]=="CONFORME (EXCEÇÃO)").sum())
        ic  = int((sub["resultado"]=="INCONFORME").sum())
        ie  = int((sub["resultado"]=="INCONFORME (EXCEÇÃO)").sum())
        tok = cf + ce
        summary.append({"Matrícula":mat,"Nome":nome,"Total":t_,"Conf.":cf,"Conf.(Exc.)":ce,
                         "Inc.":ic,"Inc.(Exc.)":ie,"Total Conf.":tok,
                         "Conform.":f"{tok/t_*100:.1f}%" if t_ else "0%"})
    st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)

    # ── DOWNLOAD
    st.markdown('<div class="section-title">⬇️ Download do Relatório</div>', unsafe_allow_html=True)
    ps = (f"_{periodo_ini.strftime('%d%m%Y')}_a_{periodo_fim.strftime('%d%m%Y')}"
          if periodo_ini and periodo_fim else "")
    excel_bytes = gerar_excel(df_res, restaurante, periodo_ini, periodo_fim)
    dl_col, _ = st.columns([1, 3])
    with dl_col:
        st.download_button("📥  Baixar Relatório Excel (.xlsx)", data=excel_bytes,
                           file_name=f"AUDITORIA_REFEICOES{ps}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    st.caption("Sistermi Auditoria de Refeições v1.2")

if __name__ == "__main__":
    main()
